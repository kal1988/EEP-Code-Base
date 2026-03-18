import os
import json
import shutil
from uuid import uuid4
from datetime import date
from typing import List, Optional
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Form, Request, logger
from fastapi.responses import JSONResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles
from sqlmodel import Session, select
from sqlalchemy import text
from jose import JWTError, jwt
from fastapi.middleware.cors import CORSMiddleware
import traceback
import logging
import openpyxl
from io import BytesIO

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

from database import create_db_and_tables, get_session
from models import Employee, Department, DepartmentUpdate, CEOAssign, Position, User, UserCreate, Token, TokenData, OrganizationSettings, ActivityLog
from auth import verify_password, get_password_hash, create_access_token, ALGORITHM, SECRET_KEY

UPLOAD_DIR = "uploads"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

app = FastAPI(title="Ethiopian Electric Power Comprehensive Employees Data Management System API")

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.error(f"Global error caught: {exc}")
    logger.error(traceback.format_exc())
    return JSONResponse(
        status_code=500,
        content={"detail": str(exc), "traceback": traceback.format_exc()},
    )

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["Authorization", "Content-Type"],
)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

async def get_current_user(token: str = Depends(oauth2_scheme), session: Session = Depends(get_session)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username)
    except JWTError:
        raise credentials_exception
    user = session.exec(select(User).where(User.username == token_data.username)).first()
    if user is None:
        raise credentials_exception
    return user

def save_upload(file: UploadFile, prefix: str = "upload") -> str:
    file_ext = file.filename.split(".")[-1]
    file_name = f"{prefix}_{uuid4()}.{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, file_name)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return f"/uploads/{file_name}"

def log_activity(session: Session, action: str, description: str, entity_type: str, actor: str, icon: str = "📋", details: Optional[str] = None):
    """Create a new activity log entry."""
    entry = ActivityLog(
        action=action,
        description=description,
        entity_type=entity_type,
        actor=actor,
        icon=icon,
        details=details
    )
    session.add(entry)
    # Don't commit here — let the caller commit once for both the main action + the log

@app.on_event("startup")
def on_startup():
    from database import engine
    create_db_and_tables()
    
    # Manual migration for activitylog, employee, department, and org tables
    try:
        # Migration logic: Individual transactions for resilience
        logger.info("Verifying database schema...")
        
        def add_column_if_not_exists(table, column, col_type):
            with engine.connect() as migration_conn:
                try:
                    # Check if column exists first to avoid unnecessary errors
                    check_sql = text(f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table}' AND column_name = '{column}';")
                    result = migration_conn.execute(check_sql).fetchone()
                    
                    if not result:
                        logger.info(f"Adding column {column} to {table}...")
                        migration_conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {col_type};"))
                        migration_conn.commit()
                    else:
                        logger.debug(f"Column {column} already exists in {table}.")
                except Exception as step_exc:
                    logger.warning(f"Failed to add column {column} to {table}: {step_exc}")

        # Activity Log
        add_column_if_not_exists("activitylog", "details", "VARCHAR")

        # Employee
        employee_cols = [
            ("phone", "VARCHAR"),
            ("date_of_birth", "DATE"),
            ("gender", "VARCHAR"),
            ("address", "VARCHAR"),
            ("national_id", "VARCHAR"),
            ("emergency_contact_name", "VARCHAR"),
            ("emergency_contact_phone", "VARCHAR"),
            ("date_of_joining", "DATE"),
            ("is_active", "BOOLEAN"),
            ("photo_url", "VARCHAR"),
            ("document_url", "VARCHAR"),
            ("department_id", "INTEGER"),
            ("position_id", "INTEGER"),
            ("company_id", "VARCHAR"),
        ]
        for c_name, c_type in employee_cols:
            add_column_if_not_exists("employee", c_name, c_type)

        # Dept and Org
        add_column_if_not_exists("department", "parent_id", "INTEGER")
        add_column_if_not_exists("department", "responsible_employee_id", "INTEGER")
        add_column_if_not_exists("organizationsettings", "ceo_employee_id", "INTEGER")
        
        # User table migrations
        add_column_if_not_exists("user", "is_superuser", "BOOLEAN")

        logger.info("Database schema verification completed.")
    except Exception as e:
        logger.error(f"Schema verification failed: {e}")
        logger.error(traceback.format_exc())

    with Session(engine) as session:
        org = session.exec(select(OrganizationSettings)).first()
        if not org:
            org = OrganizationSettings(
                name="Ethiopian Electric Power",
                subtitle="HR Management Portal"
            )
            session.add(org)
            session.commit()

@app.get("/")
def read_root():
    return {"message": "EEP HR Management System API"}

# ─── Organization ────────────────────────────────────────────────────────────

@app.get("/organization/", response_model=OrganizationSettings)
def get_organization(session: Session = Depends(get_session)):
    org = session.exec(select(OrganizationSettings)).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization settings not found")
    return org


@app.get("/ceo/", response_model=Optional[Employee])
def get_ceo(session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    org = session.exec(select(OrganizationSettings)).first()
    if not org or not org.ceo_employee_id:
        return None
    ceo = session.get(Employee, org.ceo_employee_id)
    return ceo


@app.post("/ceo/", response_model=Employee)
def assign_ceo(
    payload: CEOAssign,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    employee = session.exec(select(Employee).where(Employee.company_id == payload.company_id)).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee with that company ID not found")

    org = session.exec(select(OrganizationSettings)).first()
    if not org:
        org = OrganizationSettings()
        session.add(org)
        session.flush()

    old_ceo = session.get(Employee, org.ceo_employee_id) if org.ceo_employee_id else None
    org.ceo_employee_id = employee.id
    session.add(org)

    changes = [
        {
            "field": "CEO",
            "old": f"{old_ceo.first_name} {old_ceo.last_name}" if old_ceo else "None",
            "new": f"{employee.first_name} {employee.last_name}",
        }
    ]

    log_activity(
        session,
        "CEO Assigned",
        f"CEO assigned to {employee.first_name} {employee.last_name} by {current_user.username}",
        "settings",
        current_user.username,
        "⭐",
        details=json.dumps(changes),
    )

    session.commit()
    session.refresh(employee)
    return employee

@app.post("/organization/", response_model=OrganizationSettings)
async def update_organization(
    name: str = Form("Ethiopian Electric Power"),
    subtitle: str = Form("HR Management Portal"),
    logo: Optional[UploadFile] = File(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    try:
        org = session.exec(select(OrganizationSettings)).first()
        
        # Track changes for detailed logging
        changes_list = []
        
        if not org:
            org = OrganizationSettings(name=name, subtitle=subtitle)
            session.add(org)
            changes_list.append({"field": "Organization", "old": "None", "new": name})
        else:
            # Better comparison handling strings and None
            if (org.name or "").strip() != (name or "").strip():
                changes_list.append({"field": "Name", "old": org.name, "new": name})
            if (org.subtitle or "").strip() != (subtitle or "").strip():
                changes_list.append({"field": "Subtitle", "old": org.subtitle or "None", "new": subtitle})
            
            org.name = name
            org.subtitle = subtitle

        logo_updated = False
        if logo:
            new_logo_url = save_upload(logo, "logo")
            changes_list.append({"field": "Logo", "old": org.logo_url or "None", "new": new_logo_url})
            org.logo_url = new_logo_url
            logo_updated = True

        session.add(org)

        # Log with details
        summary_parts = []
        for c in changes_list:
            if c['field'] == "Name":
                summary_parts.append(f"name changed to '{c['new']}'")
            elif c['field'] == "Subtitle":
                summary_parts.append("subtitle updated")
            elif c['field'] == "Logo":
                summary_parts.append("logo updated")

        summary_suffix = f": {', '.join(summary_parts)}" if summary_parts else ""
        
        if len(changes_list) == 1 and changes_list[0]['field'] == "Name":
            prefix = f"settings display name field updated by {current_user.username}"
        else:
            prefix = f"Organization settings updated by {current_user.username}"
            
        description = f"{prefix}{summary_suffix}"
        
        details_json = json.dumps(changes_list) if changes_list else None
        log_activity(session, "Settings Updated", description, "settings", current_user.username, "⚙️", details=details_json)

        session.commit()
        session.refresh(org)
        return org
    except Exception as e:
        logger.error(f"Error updating organization: {e}")
        logger.error(traceback.format_exc())
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ─── User Management ─────────────────────────────────────────────────────────

@app.get("/users/", response_model=List[User])
def list_users(session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    """List all registered users (experimental — we should ideally restrict this to superusers further)."""
    users = session.exec(select(User)).all()
    return users

@app.delete("/users/{user_id}", status_code=204)
def delete_user(user_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    target_user = session.get(User, user_id)
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
        
    if target_user.username == current_user.username:
        raise HTTPException(status_code=400, detail="You cannot delete yourself")

    try:
        username_to_delete = target_user.username
        session.delete(target_user)
        
        log_activity(
            session,
            "User Deleted",
            f"Administrator user '{username_to_delete}' was removed by {current_user.username}",
            "user",
            current_user.username,
            "🗑️"
        )
        
        session.commit()
    except Exception as e:
        logger.error(f"Error deleting user: {e}")
        session.rollback()
        raise HTTPException(status_code=500, detail="Failed to delete user")
    
    return None

from pydantic import BaseModel

class PasswordUpdate(BaseModel):
    new_password: str

@app.put("/users/{user_id}/password", status_code=200)
def change_user_password(user_id: int, pwd_update: PasswordUpdate, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    target_user = session.get(User, user_id)
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    try:
        target_user.hashed_password = get_password_hash(pwd_update.new_password)
        session.add(target_user)
        
        log_activity(
            session,
            "Password Changed",
            f"Administrator {current_user.username} changed the password for user '{target_user.username}'",
            "user",
            current_user.username,
            "🔑"
        )
        
        session.commit()
    except Exception as e:
        logger.error(f"Error updating password: {e}")
        session.rollback()
        raise HTTPException(status_code=500, detail="Failed to update password")
    
    return {"message": "Password updated successfully"}

# ─── Auth ─────────────────────────────────────────────────────────────────────

@app.post("/register", response_model=User)
def register(user_data: UserCreate, session: Session = Depends(get_session)):
    existing_user = session.exec(select(User).where(User.username == user_data.username)).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already registered")
    hashed_password = get_password_hash(user_data.password)
    new_user = User(
        username=user_data.username,
        email=user_data.email,
        hashed_password=hashed_password
    )
    session.add(new_user)
    session.commit()
    session.refresh(new_user)
    return new_user

@app.post("/token", response_model=Token)
def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), session: Session = Depends(get_session)):
    user = session.exec(select(User).where(User.username == form_data.username)).first()
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token = create_access_token(data={"sub": user.username})
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/users/me", response_model=User)
def read_users_me(current_user: User = Depends(get_current_user)):
    return current_user

# ─── Activity Logs ────────────────────────────────────────────────────────────

@app.get("/activity-logs/")
def get_activity_logs(
    limit: int = 20,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    logs = session.exec(
        select(ActivityLog).order_by(ActivityLog.timestamp.desc()).limit(limit)
    ).all()
    return logs

# ─── Employees ────────────────────────────────────────────────────────────────

@app.get("/employees/", response_model=List[Employee])
def read_employees(session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    return session.exec(select(Employee)).all()

app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

@app.post("/employees/", response_model=Employee, status_code=status.HTTP_201_CREATED)
async def create_employee(
    first_name: str = Form(...),
    last_name: str = Form(...),
    email: str = Form(...),
    company_id: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    date_of_birth: Optional[date] = Form(None),
    gender: Optional[str] = Form(None),
    address: Optional[str] = Form(None),
    national_id: Optional[str] = Form(None),
    emergency_contact_name: Optional[str] = Form(None),
    emergency_contact_phone: Optional[str] = Form(None),
    department_id: Optional[int] = Form(None),
    position_id: Optional[int] = Form(None),
    photo: Optional[UploadFile] = File(None),
    document: Optional[UploadFile] = File(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    try:
        photo_url = save_upload(photo, "photo") if photo else None
        document_url = save_upload(document, "doc") if document else None

        employee = Employee(
            first_name=first_name,
            last_name=last_name,
            email=email,
            company_id=company_id,
            phone=phone,
            date_of_birth=date_of_birth,
            gender=gender,
            address=address,
            national_id=national_id,
            emergency_contact_name=emergency_contact_name,
            emergency_contact_phone=emergency_contact_phone,
            department_id=department_id,
            position_id=position_id,
            photo_url=photo_url,
            document_url=document_url
        )
        session.add(employee)

        # Look up department name for a nice log message
        dept_name = ""
        if department_id:
            dept = session.get(Department, department_id)
            if dept:
                dept_name = f" in {dept.name}"

        log_activity(
            session,
            "Employee Added",
            f"{first_name} {last_name} was onboarded{dept_name} by {current_user.username}",
            "employee",
            current_user.username,
            "👤"
        )

        session.commit()
        session.refresh(employee)
        logger.info(f"Employee created successfully: {employee.id}")
        return employee
    except Exception as e:
        logger.error(f"Error creating employee: {e}")
        logger.error(traceback.format_exc())
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/employees/{employee_id}", response_model=Employee)
def read_employee(employee_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    employee = session.get(Employee, employee_id)
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    return employee

@app.patch("/employees/{employee_id}", response_model=Employee)
async def update_employee(
    employee_id: int,
    company_id: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    address: Optional[str] = Form(None),
    emergency_contact_name: Optional[str] = Form(None),
    emergency_contact_phone: Optional[str] = Form(None),
    department_id: Optional[int] = Form(None),
    position_id: Optional[int] = Form(None),
    is_active: Optional[bool] = Form(None),
    photo: Optional[UploadFile] = File(None),
    document: Optional[UploadFile] = File(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    employee = session.get(Employee, employee_id)
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    changes_list = []
    
    if company_id is not None and company_id != employee.company_id:
        changes_list.append({"field": "Company ID", "old": employee.company_id or "None", "new": company_id})
        employee.company_id = company_id
    if phone is not None and phone != employee.phone:
        changes_list.append({"field": "Phone", "old": employee.phone or "None", "new": phone})
        employee.phone = phone
    if address is not None and address != employee.address:
        changes_list.append({"field": "Address", "old": employee.address or "None", "new": address})
        employee.address = address
    if emergency_contact_name is not None and emergency_contact_name != employee.emergency_contact_name:
        changes_list.append({"field": "Emergency Contact Name", "old": employee.emergency_contact_name or "None", "new": emergency_contact_name})
        employee.emergency_contact_name = emergency_contact_name
    if emergency_contact_phone is not None and emergency_contact_phone != employee.emergency_contact_phone:
        changes_list.append({"field": "Emergency Contact Phone", "old": employee.emergency_contact_phone or "None", "new": emergency_contact_phone})
        employee.emergency_contact_phone = emergency_contact_phone
    if department_id is not None and department_id != employee.department_id:
        old_dept = session.get(Department, employee.department_id) if employee.department_id else None
        new_dept = session.get(Department, department_id)
        changes_list.append({"field": "Department", "old": old_dept.name if old_dept else "None", "new": new_dept.name if new_dept else str(department_id)})
        employee.department_id = department_id
    if position_id is not None and position_id != employee.position_id:
        old_pos = session.get(Position, employee.position_id) if employee.position_id else None
        new_pos = session.get(Position, position_id)
        changes_list.append({"field": "Position", "old": old_pos.title if old_pos else "None", "new": new_pos.title if new_pos else str(position_id)})
        employee.position_id = position_id
    if is_active is not None and is_active != employee.is_active:
        changes_list.append({"field": "Active Status", "old": str(employee.is_active), "new": str(is_active)})
        employee.is_active = is_active
    if photo:
        changes_list.append({"field": "Photo", "old": "Updated", "new": "New Upload"})
        employee.photo_url = save_upload(photo, "photo")
    if document:
        changes_list.append({"field": "Document", "old": "Updated", "new": "New Upload"})
        employee.document_url = save_upload(document, "doc")

    session.add(employee)

    try:
        if changes_list:
            description = f"{employee.first_name} {employee.last_name}'s profile updated by {current_user.username}"
            details_json = json.dumps(changes_list)
            log_activity(
                session,
                "Employee Updated",
                description,
                "employee",
                current_user.username,
                "✏️",
                details=details_json
            )

        session.commit()
        session.refresh(employee)
        return employee
    except Exception as e:
        logger.error(f"Error updating employee: {e}")
        logger.error(traceback.format_exc())
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/employees/{employee_id}", status_code=204)
def delete_employee(employee_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    employee = session.get(Employee, employee_id)
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Log the deletion
    log_activity(
        session,
        "Employee Deleted",
        f"Employee {employee.first_name} {employee.last_name} was removed by {current_user.username}",
        "employee",
        current_user.username,
        "🗑️"
    )
    
    session.delete(employee)
    session.commit()
    return None

@app.post("/employees/bulk-upload", response_model=dict)
async def bulk_upload_employees(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents))
        ws = wb.active
        
        # Read headers
        headers = [str(cell.value).strip().lower() for cell in ws[1] if cell.value]
        
        added_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue # Skip empty rows
            
            row_dict = dict(zip(headers, row))
            first_name = str(row_dict.get("first_name", "")).strip()
            last_name = str(row_dict.get("last_name", "")).strip()
            email = str(row_dict.get("email", "")).strip()
            
            if not first_name or not last_name or not email or first_name == "None":
                continue 
                
            # Check if employee exists
            existing = session.exec(select(Employee).where(Employee.email == email)).first()
            if existing: continue 
            
            new_emp = Employee(
                first_name=first_name,
                last_name=last_name,
                email=email,
                company_id=str(row_dict.get("company_id", "")).strip() if row_dict.get("company_id") else None,
                phone=str(row_dict.get("phone", "")).strip() if row_dict.get("phone") else None,
                gender=str(row_dict.get("gender", "")).strip() if row_dict.get("gender") else None,
                national_id=str(row_dict.get("national_id", "")).strip() if row_dict.get("national_id") else None,
                address=str(row_dict.get("address", "")).strip() if row_dict.get("address") else None,
            )
            session.add(new_emp)
            added_count += 1
            
        if added_count > 0:
            log_activity(
                session,
                "Bulk Migration",
                f"Imported {added_count} employees from Excel",
                "employee",
                current_user.username,
                "📥"
            )
            session.commit()
            
        return {"message": f"Successfully imported {added_count} employees.", "count": added_count}

    except Exception as e:
        session.rollback()
        logger.error(f"Error processing Excel file: {e}")
        raise HTTPException(status_code=500, detail=f"Error reading file: {str(e)}")

# ─── Departments ──────────────────────────────────────────────────────────────

@app.get("/departments/", response_model=List[Department])
def read_departments(session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    return session.exec(select(Department)).all()

@app.post("/departments/", response_model=Department)
def create_department(department: Department, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    session.add(department)
    log_activity(
        session,
        "Business Unit Created",
        f"New business unit '{department.name}' created by {current_user.username}",
        "department",
        current_user.username,
        "🏢"
    )
    session.commit()
    session.refresh(department)
    return department


@app.patch("/departments/{department_id}", response_model=Department)
def update_department(
    department_id: int,
    payload: DepartmentUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    department = session.get(Department, department_id)
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")

    changes = []

    if payload.name is not None and payload.name != department.name:
        changes.append({"field": "Name", "old": department.name, "new": payload.name})
        department.name = payload.name

    if payload.description is not None and payload.description != department.description:
        changes.append(
            {
                "field": "Description",
                "old": department.description or "",
                "new": payload.description,
            }
        )
        department.description = payload.description

    if payload.parent_id is not None and payload.parent_id != department.parent_id:
        old_parent = (
            session.get(Department, department.parent_id)
            if department.parent_id
            else None
        )
        new_parent = session.get(Department, payload.parent_id)
        changes.append(
            {
                "field": "Parent Unit",
                "old": old_parent.name if old_parent else "Top-level",
                "new": new_parent.name if new_parent else "Top-level",
            }
        )
        department.parent_id = payload.parent_id

    if (
        payload.responsible_employee_id is not None
        and payload.responsible_employee_id != department.responsible_employee_id
    ):
        from_employee = (
            session.get(Employee, department.responsible_employee_id)
            if department.responsible_employee_id
            else None
        )
        to_employee = session.get(Employee, payload.responsible_employee_id)
        changes.append(
            {
                "field": "Responsible Person",
                "old": f"{from_employee.first_name} {from_employee.last_name}"
                if from_employee
                else "None",
                "new": f"{to_employee.first_name} {to_employee.last_name}"
                if to_employee
                else "None",
            }
        )
        department.responsible_employee_id = payload.responsible_employee_id

    session.add(department)

    if changes:
        log_activity(
            session,
            "Business Unit Updated",
            f"Business unit '{department.name}' updated by {current_user.username}",
            "department",
            current_user.username,
            "🏢",
            details=json.dumps(changes),
        )

    session.commit()
    session.refresh(department)
    return department

@app.delete("/departments/{department_id}", status_code=204)
def delete_department(department_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    dept = session.get(Department, department_id)
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    log_activity(
        session,
        "Business Unit Deleted",
        f"Business unit '{dept.name}' deleted by {current_user.username}",
        "department",
        current_user.username,
        "🗑️"
    )
    session.delete(dept)
    session.commit()

# ─── Positions ────────────────────────────────────────────────────────────────

@app.get("/positions/", response_model=List[Position])
def read_positions(session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    return session.exec(select(Position)).all()

@app.post("/positions/", response_model=Position)
def create_position(position: Position, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    session.add(position)
    log_activity(
        session,
        "Position Created",
        f"New position '{position.title}' created by {current_user.username}",
        "position",
        current_user.username,
        "💼"
    )
    session.commit()
    session.refresh(position)
    return position

@app.delete("/positions/{position_id}", status_code=204)
def delete_position(position_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)):
    pos = session.get(Position, position_id)
    if not pos:
        raise HTTPException(status_code=404, detail="Position not found")
    log_activity(
        session,
        "Position Deleted",
        f"Position '{pos.title}' deleted by {current_user.username}",
        "position",
        current_user.username,
        "🗑️"
    )
    session.delete(pos)
    session.commit()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
